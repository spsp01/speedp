from openpyxl import load_workbook
from openpyxl.drawing.image import Image


def printtest(payload):
    for i, url in enumerate(payload['responsecodes']):
        print(i)
        print(url)

def textspeed(score):
    if score >= 90:
        speedtext = 'GOOD'
    if score >= 50 and score < 90:
        speedtext = 'MEDIUM'
    if score <50:
        speedtext = 'POOR'
    return speedtext

def condition(text):
    if text == 0:
        return 'OK'
    else:
        return 'Błąd'

def createraport(payload):
    wb = load_workbook('xls/raport.xlsx') # template

    renderd = wb['Desktop'] # desktop
    titlep = ['J5','J16','J27','J38','J49']
    imagep = ['C4','C15','C26','C37','C48']
    urlp = ['J7','J18','J29','J40','J51']
    speedp = ['S7', 'S18', 'S29', 'S40', 'S51']
    speedtextp=['S8', 'S19', 'S30', 'S41', 'S52']
    responsep = ['J9','J20','J31','J42','J53']

    renderm = wb['Mobile']  # mobile
    speedtextm = ['E46', 'M46', 'U46', 'AC46', 'AK46']
    speedm = ['E45', 'M45', 'U45', 'AC45', 'AK45']
    imagem = ['C5','K5','S5','AA5','AI5']
    titlem = ['C35', 'K35', 'S35', 'AA35', 'AI35']
    urlm = ['C39', 'K39', 'S39', 'AA39', 'AI39']

    for i in range(5):
        img = Image('img/tmp/desktop/'+str(i+1)+'.jpg')
        renderd.add_image(img, imagep[i])

    for i in range(5):
        img = Image('img/tmp/mobile/' + str(i + 1) + '.jpg')
        renderm.add_image(img, imagem[i])

    for i, title in enumerate(payload['titles']):
        renderd[titlep[i]] = str(title)
        renderm[titlem[i]] =str(title)

    for i, url in enumerate(payload['urls']):
        renderd[urlp[i]] = str(url)
        renderm[urlm[i]] = str(url)

    for i, resp in enumerate(payload['responsecodes']):
        renderd[responsep[i]] = resp

    for i, speedscore in enumerate(payload['speedscores']):
        renderd[speedp[i]] = speedscore
        speedtext = textspeed(speedscore)
        renderd[speedtextp[i]] = speedtext

    for i, speedscorem in enumerate(payload['speedscoresm']):
        renderm[speedm[i]] = speedscorem
        speedtext = textspeed(speedscorem)
        renderm[speedtextm[i]] = speedtext

    summary = wb['Statystyki']  # Podsumowanie

    for i,stat in enumerate(payload['stats']):
        summary['B'+str(5+i)] = payload['stats'][i]['url']
        summary['C'+str(5+i)] = payload['stats'][i]['title']
        summary['D'+str(5+i)] = payload['stats'][i]['type']
        summary['E'+str(5+i)] = int(payload['stats'][i]['responsecode'])

        summary['F'+str(5+i)] = int(payload['stats'][i]['numberResources'])
        summary['G'+str(5+i)] = int(payload['stats'][i]['numberHosts'])
        summary['H'+str(5+i)] = int(payload['stats'][i]['totalRequestBytes'])
        summary['I'+str(5+i)] = int(payload['stats'][i]['numberStaticResources'])

        summary['J'+str(5+i)] = int(payload['stats'][i]['htmlResponseBytes'])
        summary['K'+str(5+i)] = int(payload['stats'][i]['textResponseBytes'])
        summary['L'+str(5+i)] = int(payload['stats'][i]['overTheWireResponseBytes'])
        summary['M'+str(5+i)] = int(payload['stats'][i]['cssResponseBytes'])

        summary['N'+str(5+i)] = int(payload['stats'][i]['imageResponseBytes'])
        summary['O'+str(5+i)] = int(payload['stats'][i]['javascriptResponseBytes'])
        summary['P'+str(5+i)] = int(payload['stats'][i]['otherResponseBytes'])
        summary['Q'+str(5+i)] = int(payload['stats'][i]['numberJsResources'])

        summary['R'+str(5+i)] = int(payload['stats'][i]['numberCssResources'])

    speedm = wb['Szybkość']

    for i, speed in enumerate(payload['speed']):
        speedm['B' + str(5 + i)] = payload['speed'][i]['url']
        speedm['C' + str(5 + i)] = payload['speed'][i]['title']
        speedm['D' + str(5 + i)] = payload['speed'][i]['type']
        speedm['E' + str(5 + i)] = payload['speed'][i]['responsecode']

        speedm['F' + str(5 + i)] = payload['speed'][i]['categoryFCPM']
        speedm['G' + str(5 + i)] = int(payload['speed'][i]['medianFCPM'])
        speedm['H' + str(5 + i)] = payload['speed'][i]['categoryDOM']
        speedm['I' + str(5 + i)] = int(payload['speed'][i]['medianDOM'])
        speedm['J' + str(5 + i)] = int(payload['speed'][i]['speedScore'])
        speedm['K' + str(5 + i)] = int(payload['speed'][i]['AvoidLandingPageRedirects'])
        speedm['L' + str(5 + i)] = int(payload['speed'][i]['AvoidPlugins'])
        speedm['M' + str(5 + i)] = int(payload['speed'][i]['ConfigureViewport'])
        speedm['N' + str(5 + i)] = int(payload['speed'][i]['EnableGzipCompression'])
        speedm['O' + str(5 + i)] = int(payload['speed'][i]['LeverageBrowserCaching'])
        speedm['P' + str(5 + i)] = int(payload['speed'][i]['MainResourceServerResponseTime'])
        speedm['Q' + str(5 + i)] = int(payload['speed'][i]['MinifyCss'])
        speedm['R' + str(5 + i)] = int(payload['speed'][i]['MinifyHTML'])
        speedm['S' + str(5 + i)] = int(payload['speed'][i]['MinifyJavaScript'])
        speedm['T' + str(5 + i)] = int(payload['speed'][i]['MinimizeRenderBlockingResources'])
        speedm['U' + str(5 + i)] = int(payload['speed'][i]['OptimizeImages'])
        speedm['V' + str(5 + i)] = int(payload['speed'][i]['PrioritizeVisibleContent'])

    uxm = wb['UX mobile']
    c = 0
    for i, speed in enumerate(payload['speed']):
        if speed['type'] == 'MOBILE':
            uxm['B' + str(5 + c)] = payload['speed'][i]['url']
            uxm['C' + str(5 + c)] = payload['speed'][i]['title']
            uxm['D' + str(5 + c)] = payload['speed'][i]['type']
            uxm['E' + str(5 + c)] = payload['speed'][i]['responsecode']
            uxm['F' + str(5 + c)] = condition(payload['speed'][i]['SizeContentToViewport'])
            uxm['G' + str(5 + c)] = condition(payload['speed'][i]['SizeTapTargetsAppropriately'])
            uxm['H' + str(5 + c)] = condition(payload['speed'][i]['UseLegibleFontSizes'])
            c+=1

    print(wb.sheetnames)
    wb.save(filename = 'newfile.xlsx')